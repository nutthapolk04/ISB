"""
Bulk Import API — admin/manager endpoints for Excel and CSV import.

POST /admin/import/products      — upsert ShopProduct rows by barcode
POST /admin/import/stock-receive — receive stock via ShopMovement / InventoryService
"""
from __future__ import annotations

import csv
import io
import logging
import time
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.deps import require_role
from app.core.database import get_db
from app.models.shop import MovementType, Shop, ShopCategory, ShopMovement, ShopProduct
from app.services.audit_service import create_audit_log
from sqlalchemy.exc import IntegrityError, DataError, SQLAlchemyError


def _friendly_db_error(exc: Exception) -> str:
    """Translate raw DB/ORM errors into a short Thai-friendly message so the
    import preview surfaces something an operator can act on, instead of a
    stack-trace fragment."""
    msg = str(exc)
    low = msg.lower()
    if isinstance(exc, IntegrityError) or "unique" in low or "duplicate" in low:
        if "barcode" in low:
            return "Barcode นี้มีในระบบแล้ว (ระบุ barcode ซ้ำกับสินค้าอื่น)"
        if "product_code" in low:
            return "รหัสสินค้านี้มีในระบบแล้ว"
        return "ข้อมูลซ้ำกับรายการที่มีอยู่ในระบบ"
    if isinstance(exc, DataError) or "out of range" in low or "value too long" in low:
        return "รูปแบบข้อมูลไม่ถูกต้อง (เช่น ตัวเลขเกินช่วง หรือข้อความยาวเกินไป)"
    if "not-null" in low or "null value" in low:
        return "พบช่องที่จำเป็นแต่ปล่อยว่าง"
    if isinstance(exc, SQLAlchemyError):
        return f"บันทึกไม่สำเร็จ: {msg[:200]}"
    return msg[:200]
from app.models.unit_of_measure import UnitOfMeasure
from app.models.user import User
from app.services.inventory_service import InventoryService

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Response schemas ──────────────────────────────────────────────────────────

class ImportError(BaseModel):
    row: int
    reason: str


class ProductImportResult(BaseModel):
    created: int
    updated: int
    errors: List[ImportError]


class StockImportResult(BaseModel):
    imported: int
    errors: List[ImportError]


class StoreImportResult(BaseModel):
    products: ProductImportResult
    stock: StockImportResult


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_file(filename: str, content: bytes, preferred_sheet: Optional[str] = None) -> list[dict]:
    """
    Parse uploaded file to list of row dicts.
    Supports .xlsx (via openpyxl) and .csv (via stdlib csv).

    When `preferred_sheet` is provided and the workbook contains a sheet by
    that name, that sheet is used. Otherwise falls back to the active sheet.
    Lets one xlsx ship two import targets (Products + StockReceive) without
    forcing the operator to delete or reorder sheets before upload.
    """
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    if ext == "xlsx":
        import openpyxl  # already in requirements.txt
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        if preferred_sheet and preferred_sheet in wb.sheetnames:
            ws = wb[preferred_sheet]
        else:
            ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        return result

    elif ext == "csv":
        text = content.decode("utf-8-sig")  # strip BOM if present
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Please upload a .xlsx or .csv file.",
        )


def _str(val) -> str:
    """Coerce cell value to stripped string, empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


def _float_or_none(val) -> Optional[float]:
    s = _str(val)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _int_or_none(val) -> Optional[int]:
    f = _float_or_none(val)
    return int(f) if f is not None else None


# ── Row-processing helpers (shared by individual and combined endpoints) ─────

def _run_products_rows(
    rows: list,
    shop_id: str,
    is_manager: bool,
    current_user: User,
    db: Session,
    dry_run: bool,
) -> ProductImportResult:
    """Process parsed product rows. Caller must commit or rollback after."""
    import secrets as _secrets

    created = 0
    updated = 0
    errors: List[ImportError] = []

    for idx, row in enumerate(rows, start=2):
        sp = db.begin_nested()
        try:
            row_shop_id = _str(row.get("shop_id")) or shop_id
            if not row_shop_id:
                errors.append(ImportError(row=idx, reason="ต้องระบุ shop_id (ทั้งเป็น query param หรือคอลัมน์ในไฟล์)"))
                sp.rollback()
                continue
            if is_manager and row_shop_id != current_user.shop_id:
                errors.append(ImportError(row=idx, reason="Manager นำเข้าได้เฉพาะร้านของตัวเองเท่านั้น"))
                sp.rollback()
                continue

            shop = db.query(Shop).filter(Shop.id == row_shop_id).first()
            if not shop:
                errors.append(ImportError(row=idx, reason=f"ไม่พบร้าน '{row_shop_id}' ในระบบ"))
                sp.rollback()
                continue

            name = _str(row.get("name"))
            barcode = _str(row.get("barcode"))
            price_val = _float_or_none(row.get("price"))
            cost_val = _float_or_none(row.get("cost_price"))

            if not name:
                errors.append(ImportError(row=idx, reason="ต้องระบุ 'name' (ชื่อสินค้า)"))
                sp.rollback()
                continue
            if price_val is None:
                errors.append(ImportError(row=idx, reason="'price' (ราคาขาย) ต้องเป็นตัวเลข"))
                sp.rollback()
                continue
            if cost_val is None:
                errors.append(ImportError(row=idx, reason="'cost_price' (ต้นทุน) ต้องเป็นตัวเลข"))
                sp.rollback()
                continue

            category = _str(row.get("category")) or "ทั่วไป"
            if category:
                existing_cat = (
                    db.query(ShopCategory)
                    .filter(ShopCategory.shop_id == row_shop_id, ShopCategory.name == category)
                    .first()
                )
                if not existing_cat:
                    db.add(ShopCategory(shop_id=row_shop_id, name=category))
                    db.flush()

            uom_name = _str(row.get("uom"))
            uom_id: Optional[int] = None
            if uom_name:
                uom = db.query(UnitOfMeasure).filter(
                    UnitOfMeasure.name == uom_name, UnitOfMeasure.is_active == True
                ).first()
                if uom:
                    uom_id = uom.id

            if barcode:
                existing = db.query(ShopProduct).filter(
                    ShopProduct.shop_id == row_shop_id, ShopProduct.barcode == barcode
                ).first()
            else:
                existing = db.query(ShopProduct).filter(
                    ShopProduct.shop_id == row_shop_id, ShopProduct.name == name
                ).first()

            if existing:
                old_snapshot = {
                    "name": existing.name,
                    "external_price": float(existing.external_price),
                    "internal_price": float(existing.internal_price),
                    "category": existing.category,
                }
                existing.name = name
                existing.external_price = price_val
                existing.internal_price = cost_val
                existing.category = category
                if uom_id is not None:
                    existing.uom_id = uom_id
                db.flush()
                if not dry_run:
                    create_audit_log(
                        db, entity_type="shop_product", entity_id=existing.id,
                        entity_name=existing.name, shop_id=row_shop_id,
                        action="UPDATE_PRODUCT",
                        changes={"source": "import", "old": old_snapshot,
                                 "new": {"name": name, "external_price": price_val,
                                         "internal_price": cost_val, "category": category}},
                        user_id=current_user.id,
                    )
                updated += 1
            else:
                ts_suffix = int(time.time() * 1000) % 100_000_000
                product_code = f"IMP-{ts_suffix:08d}{_secrets.token_hex(1)}"
                product = ShopProduct(
                    shop_id=row_shop_id, product_code=product_code,
                    barcode=barcode or None, name=name, category=category,
                    external_price=price_val, internal_price=cost_val,
                    uom_id=uom_id, stock=0,
                )
                db.add(product)
                db.flush()
                if not dry_run:
                    create_audit_log(
                        db, entity_type="shop_product", entity_id=product.id,
                        entity_name=name, shop_id=row_shop_id, action="create",
                        changes={"source": "import",
                                 "new": {"name": name, "barcode": barcode or None,
                                         "external_price": price_val,
                                         "internal_price": cost_val, "category": category}},
                        user_id=current_user.id,
                    )
                created += 1

            sp.commit()

        except Exception as exc:
            sp.rollback()
            logger.warning("import_products row %d error: %s", idx, exc)
            errors.append(ImportError(row=idx, reason=_friendly_db_error(exc)))

    return ProductImportResult(created=created, updated=updated, errors=errors)


def _run_combined_rows(
    rows: list,
    shop_id: str,
    is_manager: bool,
    current_user: User,
    db: Session,
    dry_run: bool,
) -> tuple[ProductImportResult, StockImportResult]:
    """Process rows that mix product columns + stock-receive columns in one sheet.

    For each row:
      1. Upsert the product (same rules as `_run_products_rows`)
      2. If the row also carries a positive `quantity`, receive that stock
         straight into the product we just upserted — no barcode lookup race
         because we already hold the ORM object.

    A single row therefore covers the common "new product + opening stock"
    case in one step. Rows that omit `quantity` still upsert the product
    (useful for catalog-only imports). Rows that omit product columns but
    carry `barcode` + `quantity` still do a barcode lookup so a pure
    stock-only sheet still works.

    Errors from the product step short-circuit the stock step for that row
    (no point receiving stock into a product that failed to create).
    """
    import secrets as _secrets

    p_created = 0
    p_updated = 0
    p_errors: List[ImportError] = []
    s_imported = 0
    s_errors: List[ImportError] = []

    for idx, row in enumerate(rows, start=2):
        sp = db.begin_nested()
        try:
            row_shop_id = _str(row.get("shop_id")) or shop_id
            if not row_shop_id:
                p_errors.append(ImportError(row=idx, reason="ต้องระบุ shop_id (ทั้งเป็น query param หรือคอลัมน์ในไฟล์)"))
                sp.rollback()
                continue
            if is_manager and row_shop_id != current_user.shop_id:
                p_errors.append(ImportError(row=idx, reason="Manager นำเข้าได้เฉพาะร้านของตัวเองเท่านั้น"))
                sp.rollback()
                continue

            shop = db.query(Shop).filter(Shop.id == row_shop_id).first()
            if not shop:
                p_errors.append(ImportError(row=idx, reason=f"ไม่พบร้าน '{row_shop_id}' ในระบบ"))
                sp.rollback()
                continue

            name = _str(row.get("name"))
            barcode = _str(row.get("barcode"))
            price_val = _float_or_none(row.get("price"))
            cost_val = _float_or_none(row.get("cost_price"))
            qty_val = _int_or_none(row.get("quantity"))

            # Decide whether this is a product-upsert row, stock-only row, or both.
            has_product_data = bool(name) and price_val is not None and cost_val is not None
            has_stock_data = qty_val is not None and qty_val > 0

            product: Optional[ShopProduct] = None

            if has_product_data:
                # ── Product upsert (mirrors _run_products_rows) ────────────
                category = _str(row.get("category")) or "ทั่วไป"
                if category:
                    existing_cat = (
                        db.query(ShopCategory)
                        .filter(ShopCategory.shop_id == row_shop_id, ShopCategory.name == category)
                        .first()
                    )
                    if not existing_cat:
                        db.add(ShopCategory(shop_id=row_shop_id, name=category))
                        db.flush()

                uom_name = _str(row.get("uom"))
                uom_id: Optional[int] = None
                if uom_name:
                    uom = db.query(UnitOfMeasure).filter(
                        UnitOfMeasure.name == uom_name, UnitOfMeasure.is_active == True
                    ).first()
                    if uom:
                        uom_id = uom.id

                if barcode:
                    existing = db.query(ShopProduct).filter(
                        ShopProduct.shop_id == row_shop_id, ShopProduct.barcode == barcode
                    ).first()
                else:
                    existing = db.query(ShopProduct).filter(
                        ShopProduct.shop_id == row_shop_id, ShopProduct.name == name
                    ).first()

                if existing:
                    old_snapshot = {
                        "name": existing.name,
                        "external_price": float(existing.external_price),
                        "internal_price": float(existing.internal_price),
                        "category": existing.category,
                    }
                    existing.name = name
                    existing.external_price = price_val
                    existing.internal_price = cost_val
                    existing.category = category
                    if uom_id is not None:
                        existing.uom_id = uom_id
                    db.flush()
                    if not dry_run:
                        create_audit_log(
                            db, entity_type="shop_product", entity_id=existing.id,
                            entity_name=existing.name, shop_id=row_shop_id,
                            action="UPDATE_PRODUCT",
                            changes={"source": "import", "old": old_snapshot,
                                     "new": {"name": name, "external_price": price_val,
                                             "internal_price": cost_val, "category": category}},
                            user_id=current_user.id,
                        )
                    p_updated += 1
                    product = existing
                else:
                    ts_suffix = int(time.time() * 1000) % 100_000_000
                    product_code = f"IMP-{ts_suffix:08d}{_secrets.token_hex(1)}"
                    new_product = ShopProduct(
                        shop_id=row_shop_id, product_code=product_code,
                        barcode=barcode or None, name=name, category=category,
                        external_price=price_val, internal_price=cost_val,
                        uom_id=uom_id, stock=0,
                    )
                    db.add(new_product)
                    db.flush()
                    if not dry_run:
                        create_audit_log(
                            db, entity_type="shop_product", entity_id=new_product.id,
                            entity_name=name, shop_id=row_shop_id, action="create",
                            changes={"source": "import",
                                     "new": {"name": name, "barcode": barcode or None,
                                             "external_price": price_val,
                                             "internal_price": cost_val, "category": category}},
                            user_id=current_user.id,
                        )
                    p_created += 1
                    product = new_product
            elif not (name or price_val is not None or cost_val is not None):
                # Pure stock-receive row (no product columns at all) — fall
                # through to the stock step, which will resolve the product
                # via barcode lookup below.
                pass
            else:
                # Partial product columns — that's a validation error, not a
                # silent skip. Tell the operator which field is missing.
                if not name:
                    p_errors.append(ImportError(row=idx, reason="ต้องระบุ 'name' (ชื่อสินค้า)"))
                elif price_val is None:
                    p_errors.append(ImportError(row=idx, reason="'price' (ราคาขาย) ต้องเป็นตัวเลข"))
                else:
                    p_errors.append(ImportError(row=idx, reason="'cost_price' (ต้นทุน) ต้องเป็นตัวเลข"))
                sp.rollback()
                continue

            # ── Stock receive (mirrors _run_stock_rows) ────────────────────
            if has_stock_data:
                if product is None and barcode:
                    product = db.query(ShopProduct).filter(
                        ShopProduct.barcode == barcode, ShopProduct.shop_id == row_shop_id
                    ).first()
                if product is None:
                    s_errors.append(ImportError(row=idx, reason="ไม่พบสินค้าสำหรับรับสต็อก — ต้องระบุ name/price/cost_price หรือ barcode ที่มีอยู่"))
                else:
                    cost_per_unit = _float_or_none(row.get("cost_per_unit")) or float(product.internal_price or 0)
                    notes = _str(row.get("notes")) or None
                    reference = _str(row.get("reference")) or None

                    InventoryService.receive_stock(
                        db=db, shop=shop, product=product, qty=qty_val,
                        cost_per_unit=cost_per_unit, reference=reference,
                        note=notes, user_id=current_user.id,
                    )
                    db.flush()
                    if not dry_run:
                        create_audit_log(
                            db, entity_type="shop_product", entity_id=product.id,
                            entity_name=product.name, shop_id=row_shop_id, action="update",
                            changes={"source": "stock_receive_import", "qty_received": qty_val,
                                     "cost_per_unit": cost_per_unit, "reference": reference, "notes": notes},
                            user_id=current_user.id,
                        )
                    s_imported += 1

            sp.commit()

        except Exception as exc:
            sp.rollback()
            logger.warning("import_combined row %d error: %s", idx, exc)
            # Attribute the failure to whichever step is more useful — if the
            # row had stock data, the stock bucket; otherwise products.
            target = s_errors if _int_or_none(row.get("quantity")) else p_errors
            target.append(ImportError(row=idx, reason=_friendly_db_error(exc)))

    return (
        ProductImportResult(created=p_created, updated=p_updated, errors=p_errors),
        StockImportResult(imported=s_imported, errors=s_errors),
    )


def _run_stock_rows(
    rows: list,
    is_manager: bool,
    current_user: User,
    db: Session,
    dry_run: bool,
) -> StockImportResult:
    """Process parsed stock-receive rows. Caller must commit or rollback after."""
    imported = 0
    errors: List[ImportError] = []

    for idx, row in enumerate(rows, start=2):
        sp = db.begin_nested()
        try:
            row_shop_id = _str(row.get("shop_id"))
            if not row_shop_id:
                errors.append(ImportError(row=idx, reason="ต้องระบุ 'shop_id'"))
                sp.rollback()
                continue
            if is_manager and row_shop_id != current_user.shop_id:
                errors.append(ImportError(row=idx, reason="Manager รับสต็อกได้เฉพาะร้านของตัวเองเท่านั้น"))
                sp.rollback()
                continue

            shop = db.query(Shop).filter(Shop.id == row_shop_id).first()
            if not shop:
                errors.append(ImportError(row=idx, reason=f"ไม่พบร้าน '{row_shop_id}' ในระบบ"))
                sp.rollback()
                continue

            qty = _int_or_none(row.get("quantity"))
            if qty is None or qty <= 0:
                errors.append(ImportError(row=idx, reason="'quantity' ต้องเป็นจำนวนเต็มที่มากกว่า 0"))
                sp.rollback()
                continue

            product: Optional[ShopProduct] = None
            pid_raw = _str(row.get("product_id"))
            barcode_raw = _str(row.get("barcode"))
            if pid_raw:
                pid = _int_or_none(pid_raw)
                if pid is not None:
                    product = db.query(ShopProduct).filter(
                        ShopProduct.id == pid, ShopProduct.shop_id == row_shop_id
                    ).first()
            if product is None and barcode_raw:
                product = db.query(ShopProduct).filter(
                    ShopProduct.barcode == barcode_raw, ShopProduct.shop_id == row_shop_id
                ).first()

            if product is None:
                errors.append(ImportError(row=idx, reason="ไม่พบสินค้า — กรุณาระบุ product_id หรือ barcode ที่ถูกต้อง"))
                sp.rollback()
                continue

            cost_per_unit = _float_or_none(row.get("cost_per_unit")) or float(product.internal_price or 0)
            notes = _str(row.get("notes")) or None
            reference = _str(row.get("reference")) or None

            InventoryService.receive_stock(
                db=db, shop=shop, product=product, qty=qty,
                cost_per_unit=cost_per_unit, reference=reference,
                note=notes, user_id=current_user.id,
            )
            db.flush()
            if not dry_run:
                create_audit_log(
                    db, entity_type="shop_product", entity_id=product.id,
                    entity_name=product.name, shop_id=row_shop_id, action="update",
                    changes={"source": "stock_receive_import", "qty_received": qty,
                             "cost_per_unit": cost_per_unit, "reference": reference, "notes": notes},
                    user_id=current_user.id,
                )
            imported += 1
            sp.commit()

        except Exception as exc:
            sp.rollback()
            logger.warning("import_stock_receive row %d error: %s", idx, exc)
            errors.append(ImportError(row=idx, reason=_friendly_db_error(exc)))

    return StockImportResult(imported=imported, errors=errors)


# ── POST /admin/import/products ───────────────────────────────────────────────

@router.post("/products", response_model=ProductImportResult)
async def import_products(
    file: UploadFile = File(...),
    shop_id: str = "",
    dry_run: bool = Query(False, description="Validate only — roll back instead of committing."),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "manager")),
):
    """Bulk upsert ShopProduct rows from an Excel or CSV file."""
    is_manager = current_user.role == "manager" and not current_user.is_superuser
    if is_manager:
        if not current_user.shop_id:
            raise HTTPException(status_code=403, detail="Manager has no shop assignment")
        if shop_id and shop_id != current_user.shop_id:
            raise HTTPException(status_code=403, detail="Manager can only import into their own shop")
        shop_id = current_user.shop_id

    content = await file.read()
    filename = file.filename or ""
    try:
        rows = _parse_file(filename, content)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}") from exc

    result = _run_products_rows(rows, shop_id, is_manager, current_user, db, dry_run)
    if dry_run:
        db.rollback()
    else:
        db.commit()
    return result


# ── POST /admin/import/stock-receive ─────────────────────────────────────────

@router.post("/stock-receive", response_model=StockImportResult)
async def import_stock_receive(
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Validate only — roll back instead of committing."),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "manager")),
):
    """Bulk stock receive from an Excel or CSV file."""
    is_manager = current_user.role == "manager" and not current_user.is_superuser
    if is_manager and not current_user.shop_id:
        raise HTTPException(status_code=403, detail="Manager has no shop assignment")

    content = await file.read()
    filename = file.filename or ""
    try:
        rows = _parse_file(filename, content, preferred_sheet="StockReceive")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}") from exc

    result = _run_stock_rows(rows, is_manager, current_user, db, dry_run)
    if dry_run:
        db.rollback()
    else:
        db.commit()
    return result


# ── POST /admin/import/store (combined: Products + StockReceive sheets) ───────

@router.post("/store", response_model=StoreImportResult)
async def import_store(
    file: UploadFile = File(...),
    shop_id: str = "",
    dry_run: bool = Query(False, description="Validate only — roll back instead of committing."),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "manager")),
):
    """Import a unified single-sheet xlsx where each row carries both product
    fields and an optional stock-receive (`quantity`, `cost_per_unit`, …).

    Operators told us they didn't want to flip between two sheets to onboard
    a shop — one row, one product, one opening stock entry. The combined
    handler upserts the product first, then receives stock straight into it
    when `quantity > 0`. Catalog-only rows (no quantity) and stock-only rows
    (no product columns, just barcode + quantity) both still work.

    Falls back to legacy two-sheet workbooks (Products + StockReceive) so
    existing files keep working — picked up automatically when the sheet
    detection finds those tab names.
    """
    is_manager = current_user.role == "manager" and not current_user.is_superuser
    if is_manager:
        if not current_user.shop_id:
            raise HTTPException(status_code=403, detail="Manager has no shop assignment")
        if shop_id and shop_id != current_user.shop_id:
            raise HTTPException(status_code=403, detail="Manager can only import into their own shop")
        shop_id = current_user.shop_id

    content = await file.read()
    filename = file.filename or ""

    # Legacy detection: if both Products + StockReceive sheets are present,
    # fall back to the original split-sheet flow so old template files keep
    # working without forcing operators to migrate.
    legacy_products: list = []
    legacy_stock: list = []
    try:
        legacy_products = _parse_file(filename, content, preferred_sheet="Products")
    except Exception:
        legacy_products = []
    try:
        legacy_stock = _parse_file(filename, content, preferred_sheet="StockReceive")
    except Exception:
        legacy_stock = []

    if legacy_products and legacy_stock:
        products_result = _run_products_rows(legacy_products, shop_id, is_manager, current_user, db, dry_run)
        stock_result = _run_stock_rows(legacy_stock, is_manager, current_user, db, dry_run)
    else:
        # Unified single-sheet path — read the active/first sheet (preferred=None).
        try:
            rows = _parse_file(filename, content, preferred_sheet=None)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}") from exc

        products_result, stock_result = _run_combined_rows(
            rows, shop_id, is_manager, current_user, db, dry_run
        )

    if dry_run:
        db.rollback()
    else:
        db.commit()
    return StoreImportResult(products=products_result, stock=stock_result)


# ── Template download ────────────────────────────────────────────────────────

_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _write_template_sheet(ws, headers: list[str], sample_rows: list[list]) -> None:
    """Populate `ws` with a styled header row + sample data rows."""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = max(14, len(header) + 4)

    for row_offset, row in enumerate(sample_rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_offset, column=col_idx, value=val)

    ws.freeze_panes = "A2"


@router.get("/template")
def download_import_template(
    shop_id: str = "",
    current_user: User = Depends(require_role("admin", "manager")),
    db: Session = Depends(get_db),
):
    """Download the unified single-sheet import template.

    One row = one product + (optional) its opening stock. Operators kept
    asking why they had to flip between Products and StockReceive tabs —
    this template lets them onboard a shop in one place.

    Columns:
      name, barcode, price, cost_price, category, uom, shop_id,
      quantity, cost_per_unit, notes, reference

    Leave `quantity` blank to import the product without receiving any
    stock yet. Leave the product columns blank and fill `barcode` +
    `quantity` to receive stock into an existing product.

    For canteen shops we only emit the catalog columns (no quantity /
    cost_per_unit / notes / reference) because canteens don't track
    per-SKU stock — having those columns in their template would just
    invite confused data entry.
    """
    import openpyxl

    # Resolve shop → module so we can pick relevant sample rows. Fall back to
    # bookstore samples when no shop_id is provided or the shop isn't found.
    module = "store"
    sample_shop_id = "bookstore"
    if shop_id:
        shop = db.query(Shop).filter(Shop.id == shop_id).first()
        if shop:
            module = (shop.module or "store").lower()
            sample_shop_id = shop.id

    wb = openpyxl.Workbook()
    sheet = wb.active

    if module == "canteen":
        # Catalog-only for canteens — quantity columns omitted entirely.
        sheet.title = "Menu"
        _write_template_sheet(
            sheet,
            ["name", "barcode", "price", "cost_price", "category", "uom", "shop_id"],
            [
                ["ข้าวกะเพราหมูสับ", "CT001001", 45, 28, "อาหารจานหลัก", "จาน", sample_shop_id],
                ["น้ำส้มคั้น", "CT001002", 20, 10, "เครื่องดื่ม", "แก้ว", sample_shop_id],
            ],
        )
    else:
        sheet.title = "Store"
        _write_template_sheet(
            sheet,
            [
                "name", "barcode", "price", "cost_price", "category", "uom", "shop_id",
                "quantity", "cost_per_unit", "notes", "reference",
            ],
            [
                # Row 1: new product + opening stock in one go.
                ["หนังสือคณิตศาสตร์ ม.1", "BK001001", 120, 70, "หนังสือเรียน", "เล่ม", sample_shop_id,
                 50, 65, "รับเข้าจาก supplier A", "PO-2026-001"],
                # Row 2: another combined entry — different category.
                ["สมุดบันทึก A4 80 แผ่น", "BK001002", 35, 20, "เครื่องเขียน", "เล่ม", sample_shop_id,
                 100, 18, "รับเข้าประจำเดือน", "PO-2026-002"],
                # Row 3: catalog-only — no stock receive this time.
                ["ดินสอ HB กล่อง 12 แท่ง", "BK001003", 60, 35, "เครื่องเขียน", "กล่อง", sample_shop_id,
                 None, None, None, None],
            ],
        )

    buf = io.BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="import_template.xlsx"'},
    )
